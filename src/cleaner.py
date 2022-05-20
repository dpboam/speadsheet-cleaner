from openpyxl import Workbook,load_workbook,utils

def sepFirstLastName(wbIn,wbOut,col,row):
    workbook = load_workbook(filename=wbIn)
    ws = workbook.active

    colint = utils.column_index_from_string(col)
    nextCol = utils.get_column_letter(colint + 1)
    ws.insert_cols(colint + 1)

    ws[col + str(row)]  = "First Name"
    ws[nextCol + str(row)] = "Last Name"

    for i in range(row + 1,len(ws[col]) + 1):
        name = col + str(i)
        lastName = nextCol + str(i)
        ws[lastName] = ws[name].value[ws[name].value.index(" ") + 1:]
        ws[name] = ws[name].value[:ws[name].value.index(" ")]


    workbook.save(wbOut)

sepFirstLastName("Leeds 2023 Staff List.xlsx","test2.xlsx","A",3)