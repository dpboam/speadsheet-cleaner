from openpyxl import load_workbook
from .util import clean_row

def load_sheet(wbIn, sheet, skip_rows=None):
    wb = load_workbook(filename=wbIn, read_only=True)
    skip_rows = skip_rows + 1 if skip_rows else None
    rows = wb[sheet].iter_rows(min_row=skip_rows)
    header = [h.value for h in next(rows)]
    data = [clean_row(dict(zip(header, [c.value for c in r]))) for r in rows]
    return data
